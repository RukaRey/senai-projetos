import openpyxl

workbook = openpyxl.load_workbook('cadastro.xlsx')
sheet = workbook.active
linhaAtual=3

def lerInfo():
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=6):
        nome, cpf, telefone, email, address, birth = [cell.value for cell in row]
        print(f'Nome: {nome}, CPF: {cpf}, Telefone: {telefone}, E-mail: {email}, Endereço: {address}, Data de Nascimento: {birth}')

def addRow():
    global linhaAtual
    linhaStr=str(linhaAtual)

    cellAtual="A"+linhaStr
    nome = str(input("Nome: "))
    sheet[cellAtual] = nome

    cellAtual="B"+linhaStr
    cpf = input("CPF: ")
    sheet[cellAtual] = cpf

    cellAtual="C"+linhaStr
    tell = str(input("Telefone: "))
    sheet[cellAtual] = tell

    cellAtual="D"+linhaStr
    email = str(input("E-mail: "))
    sheet[cellAtual] = email

    cellAtual="E"+linhaStr
    adrs = str(input("Endereço: "))
    sheet[cellAtual] = adrs

    cellAtual="F"+linhaStr
    brth = str(input("Data de nascimento: "))
    sheet[cellAtual] = brth
    workbook.save('cadastro.xlsx')
    linhaAtual+=1

def UpdtRow():
    cpf = input("CPF da pessoa: ")

    workbook = openpyxl.load_workbook('cadastro.xlsx')
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        cpf_na_planilha = str(row[0].value).strip()

        cpf_inserido = cpf.strip()

        if cpf_na_planilha == cpf_inserido:
            print("CPF encontrado!")

            linhaAtual = row[0].row

            nome = input("Novo nome: ")
            tell = input("Novo telefone: ")
            email = input("Novo e-mail: ")
            adrs = input("Novo endereço: ")
            brth = input("Nova data de nascimento: ")

            sheet.cell(row=linhaAtual, column=1, value=nome)
            sheet.cell(row=linhaAtual, column=3, value=tell)
            sheet.cell(row=linhaAtual, column=4, value=email)
            sheet.cell(row=linhaAtual, column=5, value=adrs)
            sheet.cell(row=linhaAtual, column=6, value=brth)

            print("Registro atualizado!")
            break
    else:
        print("CPF não encontrado.")
    workbook.save('cadastro.xlsx')

def remove_non_numeric_chars(text):
    return ''.join(filter(str.isdigit, text))

def DelRow():
    cpf = input("CPF da pessoa a ser removida: ")
    cpf = remove_non_numeric_chars(cpf)

    # Abra o arquivo Excel
    workbook = openpyxl.load_workbook('cadastro.xlsx')

    # Acesse a planilha padrão (Sheet 1)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        cpf_na_planilha = remove_non_numeric_chars(str(row[0].value))

        if cpf_na_planilha == cpf:
            linhaAtual = row[0].row
            sheet.delete_rows(linhaAtual)
            print("Registro removido!")
            break
    else:
        print("CPF não encontrado.")

    workbook.save('cadastro.xlsx')




while True:
    dell=int(input("O que deseja fazer? \n(1)Adicionar pessoa\n(2)Remover pessoa\n(3)Mostrar pessoas\n(4)Atualizar pessoa\n(5)Sair"))
    match dell:
        case 1:
            addRow()
        case 2:
            DelRow()
        case 3:
            lerInfo()
        case 4:
            UpdtRow()
        case 5:
            print("Adeus.")
            break
        case _:
            pass